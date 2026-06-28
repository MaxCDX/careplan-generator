from datetime import date, datetime
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.external_orders.adapters.base import BaseIntakeAdapter
from app.external_orders.adapters.utils import append_note_if_present, build_clinical_notes, join_nonblank_parts
from app.external_orders.errors import ExternalOrderInputError
from app.orders.schemas import OrderCreate


class EnterpriseSpreadsheetAdapter(BaseIntakeAdapter):
    """Normalize Enterprise Spreadsheet Partner workbooks into OrderCreate."""

    def parse(self, payload: dict):
        """Parse the first worksheet's header row and first data row into a dict."""
        file_path = payload.get("file_path")
        if not file_path:
            raise ExternalOrderInputError("Enterprise spreadsheet payload must include file_path.")

        try:
            workbook = load_workbook(file_path, data_only=True)
        except (OSError, InvalidFileException, BadZipFile) as exc:
            raise ExternalOrderInputError("Invalid enterprise spreadsheet payload.") from exc
        worksheet = workbook.worksheets[0]

        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers or not any(headers):
            raise ExternalOrderInputError("Enterprise spreadsheet must include headers.")

        data_row = next(rows, None)
        if not data_row:
            raise ExternalOrderInputError("Enterprise spreadsheet must include one data row.")

        cleaned_headers = [str(header).strip() if header is not None else "" for header in headers]
        if not any(cleaned_headers):
            raise ExternalOrderInputError("Enterprise spreadsheet must include headers.")

        return dict(zip(cleaned_headers, data_row))

    def transform(self, parsed) -> OrderCreate:
        """Transform a parsed Enterprise spreadsheet row into OrderCreate."""
        name_parts = [
            parsed.get("First Name", ""),
            parsed.get("Last Name", ""),
        ]
        patient_name = join_nonblank_parts(name_parts)

        clinical_note_parts = []

        notes = parsed.get("Notes", "")
        if notes:
            clinical_note_parts.append(str(notes).strip())

        clinical_note_parts.append("Source system: Enterprise Spreadsheet Partner")

        secondary_diagnoses = parsed.get("Secondary Diagnoses", "")
        if secondary_diagnoses:
            clinical_note_parts.append(f"Secondary diagnoses: {secondary_diagnoses}")

        ndc = parsed.get("NDC", "")
        append_note_if_present(clinical_note_parts, "NDC", ndc)

        dose = parsed.get("Dose", "")
        frequency = parsed.get("Frequency", "")
        dose_frequency_parts = [str(part).strip() for part in [dose, frequency] if part and str(part).strip()]
        if dose_frequency_parts:
            clinical_note_parts.append(f"Dose/frequency: {'; '.join(dose_frequency_parts)}")

        allergies = parsed.get("Allergies", "")
        append_note_if_present(clinical_note_parts, "Allergies", allergies)

        current_medications = parsed.get("Current Medications", "")
        append_note_if_present(clinical_note_parts, "Current medications", current_medications)

        weight = parsed.get("Weight", "")
        if weight:
            clinical_note_parts.append(f"Weight: {weight}")

        clinical_notes = build_clinical_notes(clinical_note_parts)

        return OrderCreate(
            patient_name=patient_name,
            patient_dob=_format_spreadsheet_date(parsed.get("Date of Birth")),
            mrn=str(parsed.get("Patient ID", "") or ""),
            provider_name=parsed.get("Provider", ""),
            provider_npi=str(parsed.get("Provider NPI", "") or ""),
            diagnosis=parsed.get("Primary Diagnosis", ""),
            medication=parsed.get("Medication", ""),
            clinical_notes=clinical_notes,
        )


def _format_spreadsheet_date(value) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    cleaned = str(value).strip()
    if not cleaned:
        return None

    return cleaned
